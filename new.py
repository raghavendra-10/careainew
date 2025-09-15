from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = "demo.xlsx"
wb = load_workbook(file_path)

# Define a yellow fill style
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Go through all sheets and highlight empty cells
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = highlight_fill

# Save as a new file so original is safe
output_path = "RFP_Response_Marked.xlsx"
wb.save(output_path)

print(f"Marked file saved as {output_path}")
